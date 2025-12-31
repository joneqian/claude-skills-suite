#!/usr/bin/env python3
"""
RAG FAQ Generator - Standardization Script
Processes Claude-expanded FAQ files and applies standardization:
- Auto-generates keywords using jieba TF-IDF
- Applies professional Excel formatting
- Performs final quality checks
"""

import sys
import json
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import re
from collections import defaultdict

# Import required libraries
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from docx import Document
    import PyPDF2
    import jieba
    import jieba.analyse
except ImportError as e:
    print(f"Error: Missing required library. Please run: pip install -r requirements.txt")
    print(f"Details: {e}")
    sys.exit(1)


class RAGFAQGenerator:
    """Generates RAG-optimized FAQ files from various sources."""

    def __init__(self, input_file: str, recommendations: Optional[Dict] = None):
        self.input_file = Path(input_file)
        self.recommendations = recommendations or {}
        self.content = None
        self.rag_data = []

    def load_document(self) -> bool:
        """Load source document."""
        if not self.input_file.exists():
            print(f"Error: File not found: {self.input_file}")
            return False

        ext = self.input_file.suffix.lower()

        try:
            if ext == '.xlsx':
                self.content = pd.read_excel(self.input_file)
            elif ext == '.docx':
                self.content = self._load_word()
            elif ext == '.pdf':
                self.content = self._load_pdf()
            elif ext == '.txt':
                self.content = self._load_text()
            else:
                print(f"Error: Unsupported file format: {ext}")
                return False

            print(f"Successfully loaded: {self.input_file}")
            return True
        except Exception as e:
            print(f"Error loading file: {e}")
            return False

    def _load_word(self) -> List[str]:
        """Load Word document."""
        doc = Document(self.input_file)
        return [p.text.strip() for p in doc.paragraphs if p.text.strip()]

    def _load_pdf(self) -> List[str]:
        """Load PDF file."""
        with open(self.input_file, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text_content = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
            return text_content

    def _load_text(self) -> List[str]:
        """Load text file."""
        with open(self.input_file, 'r', encoding='utf-8') as file:
            return [line.strip() for line in file.readlines() if line.strip()]

    def process(self) -> bool:
        """Process content and generate RAG-optimized data."""
        if isinstance(self.content, pd.DataFrame):
            return self._process_structured()
        else:
            return self._process_unstructured()

    def _process_structured(self) -> bool:
        """Process structured Excel data (expects Claude-expanded format)."""
        df = self.content

        # Find column mappings
        column_map = self._map_columns(df)

        if '问题' not in column_map:
            print("Error: Cannot find question column in the data")
            return False

        print(f"\nProcessing {len(df)} FAQ entries...")

        # Process each row
        for idx, row in df.iterrows():
            question = str(row[column_map['问题']]) if '问题' in column_map else ''
            answer = str(row[column_map['回答']]) if '回答' in column_map else ''
            category = str(row[column_map['分类']]) if '分类' in column_map else ''
            keywords = str(row[column_map['关键词']]
                           ) if '关键词' in column_map else ''

            # Skip empty questions
            if not question or question == 'nan':
                continue

            # Clean and process
            question = self._clean_text(question)
            answer = self._clean_text(answer)

            # Generate category if missing (but Claude should have provided it)
            if not category or category == 'nan':
                category = self._generate_category(question, answer)
                print(f"  Warning: Auto-generated category for Q: {question[:30]}...")

            # Always generate keywords (this is the script's main job)
            # Even if keywords exist, regenerate for consistency
            keywords = self._generate_keywords(question, answer)

            # Add to RAG data
            self.rag_data.append({
                '分类': category,
                '问题': question,
                '回答': answer,
                '关键词': keywords
            })

        print(f"✓ Processed {len(self.rag_data)} valid FAQ entries")
        return True

    def _process_unstructured(self) -> bool:
        """Process unstructured text data."""
        # Extract Q&A pairs
        qa_pairs = self._extract_qa_pairs()

        if not qa_pairs:
            print("Error: Could not extract Q&A pairs from document")
            return False

        # Process each pair
        for pair in qa_pairs:
            question = self._clean_text(pair.get('问题', ''))
            answer = self._clean_text(pair.get('回答', ''))

            if not question:
                continue

            category = self._generate_category(question, answer)
            keywords = self._generate_keywords(question, answer)

            self.rag_data.append({
                '分类': category,
                '问题': question,
                '回答': answer,
                '关键词': keywords
            })

        return True

    def _map_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """Map DataFrame columns to standard fields."""
        column_map = {}

        mapping_patterns = {
            '问题': ['问题', 'question', 'q', '提问'],
            '回答': ['回答', 'answer', 'a', '答案', '解答'],
            '分类': ['分类', 'category', '类别', '类型'],
            '关键词': ['关键词', 'keywords', 'tags', '标签']
        }

        for field, patterns in mapping_patterns.items():
            for col in df.columns:
                col_lower = str(col).lower()
                for pattern in patterns:
                    if pattern in col_lower:
                        column_map[field] = col
                        break
                if field in column_map:
                    break

        return column_map

    def _clean_text(self, text: str) -> str:
        """Clean and normalize text."""
        if not text or text == 'nan':
            return ''

        # Remove extra whitespace
        text = ' '.join(text.split())

        # Normalize punctuation
        text = text.replace('\u3000', ' ')  # Full-width space
        text = text.replace('\xa0', ' ')     # Non-breaking space

        return text.strip()

    def _generate_category(self, question: str, answer: str) -> str:
        """Generate category based on content."""
        # Combine question and answer for analysis
        text = f"{question} {answer}"

        # Common category keywords (can be extended)
        category_keywords = {
            '账户管理': ['账户', '账号', '登录', '注册', '密码', '用户名', '个人信息'],
            '支付问题': ['支付', '付款', '退款', '价格', '费用', '充值', '余额'],
            '产品功能': ['功能', '使用', '如何', '怎么', '操作', '设置'],
            '技术支持': ['错误', '故障', '问题', '无法', '不能', 'bug', '异常'],
            '订单物流': ['订单', '物流', '配送', '发货', '快递', '收货'],
            '售后服务': ['退货', '换货', '保修', '维修', '售后', '客服'],
            '安全隐私': ['安全', '隐私', '保护', '权限', '授权', '认证']
        }

        # Count keyword matches for each category
        category_scores = defaultdict(int)

        for category, keywords in category_keywords.items():
            for keyword in keywords:
                if keyword in text:
                    category_scores[category] += 1

        # Return category with highest score, or default
        if category_scores:
            return max(category_scores, key=category_scores.get)
        else:
            return '其他'

    def _generate_keywords(self, question: str, answer: str) -> str:
        """Generate keywords using jieba."""
        # Combine question and answer
        text = f"{question} {answer}"

        try:
            # Extract keywords using TF-IDF
            keywords = jieba.analyse.extract_tags(
                text, topK=5, withWeight=False)

            # Filter out very short keywords
            keywords = [kw for kw in keywords if len(kw) >= 2]

            return ','.join(keywords[:5])  # Top 5 keywords
        except Exception as e:
            print(f"Warning: Could not generate keywords: {e}")
            # Fallback: extract nouns manually
            return self._extract_simple_keywords(text)

    def _extract_simple_keywords(self, text: str) -> str:
        """Simple keyword extraction fallback."""
        # Use jieba for word segmentation
        words = jieba.cut(text)

        # Filter by length
        keywords = [w for w in words if len(w) >= 2 and len(w) <= 5]

        # Remove duplicates while preserving order
        seen = set()
        unique_keywords = []
        for kw in keywords:
            if kw not in seen:
                seen.add(kw)
                unique_keywords.append(kw)

        return ','.join(unique_keywords[:5])

    def _extract_qa_pairs(self) -> List[Dict[str, str]]:
        """Extract Q&A pairs from unstructured content."""
        qa_pairs = []
        content_str = '\n'.join(self.content) if isinstance(
            self.content, list) else str(self.content)

        # Multiple patterns to match different Q&A formats
        patterns = [
            # Format: Q: ... A: ...
            r'[Qq问][::\uff1a]\s*(.+?)\s*[Aa答回][::\uff1a]\s*(.+?)(?=[Qq问][::\uff1a]|\Z)',
            # Format: 1. Question \n Answer
            r'(\d+[\.\uff0e])\s*(.+?)\n\s*(.+?)(?=\d+[\.\uff0e]|\Z)',
            # Format: Question? \n Answer
            r'(.+?[??\uff1f])\s*\n\s*(.+?)(?=.+?[??\uff1f]|\Z)',
        ]

        for pattern in patterns:
            matches = re.finditer(pattern, content_str, re.DOTALL)
            for match in matches:
                groups = match.groups()
                if len(groups) >= 2:
                    # Extract question and answer (last two groups)
                    question = groups[-2].strip()
                    answer = groups[-1].strip()

                    # Clean numbering if present
                    question = re.sub(r'^\d+[\.\uff0e]\s*', '', question)

                    if question and answer and len(question) > 3:
                        qa_pairs.append({
                            '问题': question,
                            '回答': answer
                        })

        return qa_pairs

    def remove_duplicates(self):
        """Remove duplicate questions."""
        seen_questions = set()
        unique_data = []

        for item in self.rag_data:
            question_normalized = item['问题'].lower().strip()

            if question_normalized not in seen_questions:
                seen_questions.add(question_normalized)
                unique_data.append(item)

        removed_count = len(self.rag_data) - len(unique_data)
        if removed_count > 0:
            print(f"Removed {removed_count} duplicate questions")

        self.rag_data = unique_data

    def merge_similar(self, threshold: float = 0.85):
        """Merge similar questions (optional)."""
        # This is a placeholder for more advanced similarity merging
        # Can be implemented based on requirements
        pass

    def save_to_excel(self, output_file: str) -> bool:
        """Save processed data to Excel file with formatting."""
        if not self.rag_data:
            print("Error: No data to save")
            return False

        try:
            # Create DataFrame
            df = pd.DataFrame(self.rag_data)

            # Ensure column order
            df = df[['分类', '问题', '回答', '关键词']]

            # Save to Excel
            df.to_excel(output_file, index=False, engine='openpyxl')

            # Apply formatting
            self._apply_excel_formatting(output_file, df)

            print(f"\nSuccessfully generated RAG FAQ file: {output_file}")
            print(f"Total entries: {len(df)}")
            print(f"Categories: {df['分类'].nunique()}")

            # Print category distribution
            print("\nCategory distribution:")
            category_counts = df['分类'].value_counts()
            for category, count in category_counts.items():
                print(f"  {category}: {count}")

            return True
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False

    def _apply_excel_formatting(self, output_file: str, df: pd.DataFrame):
        """Apply formatting to match reference style."""
        # Load the workbook
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # Set sheet name
        ws.title = "知识库"

        # Define styles
        # Header style: bold, white text, blue background, centered
        header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Data style: normal text, top aligned, wrap text
        data_font = Font(name='宋体', size=11)
        data_alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=True)

        # Apply header formatting (row 1)
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Apply data formatting (rows 2 onwards)
        for row in range(2, len(df) + 2):
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = data_font
                cell.alignment = data_alignment

        # Set column widths to match reference
        ws.column_dimensions['A'].width = 15  # 分类
        ws.column_dimensions['B'].width = 40  # 问题
        ws.column_dimensions['C'].width = 80  # 回答
        ws.column_dimensions['D'].width = 25  # 关键词

        # Auto-adjust row heights based on content
        for row in range(2, len(df) + 2):
            # Calculate approximate row height based on content length
            # Get the longest content in the row
            max_lines = 1
            for col in ['A', 'B', 'C', 'D']:
                cell_value = str(ws[f'{col}{row}'].value)
                # Estimate lines based on column width
                if col == 'A':  # 分类
                    lines = max(1, len(cell_value) // 10)
                elif col == 'B':  # 问题
                    lines = max(1, len(cell_value) // 30)
                elif col == 'C':  # 回答
                    lines = max(1, len(cell_value) // 60)
                else:  # 关键词
                    lines = max(1, len(cell_value) // 20)
                max_lines = max(max_lines, lines)

            # Set row height (each line is approximately 17 points)
            ws.row_dimensions[row].height = max(17, min(max_lines * 17, 150))

        # Save the formatted workbook
        wb.save(output_file)


def main():
    """Main entry point."""
    if len(sys.argv) < 3:
        print("="*80)
        print("RAG FAQ Standardization Script")
        print("="*80)
        print("\nUsage: python generate_rag_faq.py <expanded_file> <output_file>")
        print("\nThis script expects a Claude-expanded FAQ file with columns:")
        print("  - 分类 (Category) - Required")
        print("  - 问题 (Question) - Required")
        print("  - 回答 (Answer) - Required")
        print("  - 关键词 (Keywords) - Optional (will be auto-generated)")
        print("\nExample:")
        print("  python generate_rag_faq.py FAQ_expanded.xlsx FAQ_RAG_优化版.xlsx")
        print("\nWhat this script does:")
        print("  ✓ Auto-generates keywords using jieba TF-IDF")
        print("  ✓ Applies professional Excel formatting")
        print("  ✓ Performs final duplicate check")
        print("  ✓ Creates RAG-optimized knowledge base")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    print("="*80)
    print("RAG FAQ Standardization")
    print("="*80)

    # Generate RAG FAQ
    generator = RAGFAQGenerator(input_file)

    print("\n[1/5] Loading Claude-expanded file...")
    if not generator.load_document():
        sys.exit(1)

    print("\n[2/5] Processing content and generating keywords...")
    if not generator.process():
        sys.exit(1)

    print("\n[3/5] Checking for duplicates...")
    generator.remove_duplicates()

    print("\n[4/5] Applying formatting and saving to Excel...")
    if not generator.save_to_excel(output_file):
        sys.exit(1)

    print("\n" + "="*80)
    print("✓ RAG FAQ Standardization Complete!")
    print("="*80)
    print(f"\nOutput file: {output_file}")
    print("Ready for RAG system integration!")


if __name__ == '__main__':
    main()
